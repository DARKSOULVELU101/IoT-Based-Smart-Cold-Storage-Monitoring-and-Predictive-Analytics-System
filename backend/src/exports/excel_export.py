import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, desc
from src.database.models import SensorReading, Alert, Analytics, Report, Device


class ExcelExportService:
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def __init__(self, db: AsyncSession):
        self.db = db

    def _apply_header_style(self, worksheet, headers):
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER

    def _auto_column_widths(self, worksheet):
        for col in range(1, worksheet.max_column + 1):
            max_length = 0
            for row in range(1, worksheet.max_row + 1):
                val = worksheet.cell(row=row, column=col).value
                if val is not None:
                    max_length = max(max_length, len(str(val)))
            worksheet.column_dimensions[get_column_letter(col)].width = min(max_length + 4, 40)

    def _apply_borders(self, worksheet):
        for row in range(2, worksheet.max_row + 1):
            for col in range(1, worksheet.max_column + 1):
                worksheet.cell(row=row, column=col).border = self.THIN_BORDER

    async def export_sensor_data(self, device_id=None, start_date=None, end_date=None, module_type=None):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sensor Data"

        headers = [
            "Device ID", "Module Type", "Temperature (C)", "Humidity (%)", "Door Open",
            "Door Open (s)", "Power Available", "Gas Level", "Compressor Current (A)",
            "Compressor On", "Vibration", "Current (A)", "Voltage (V)", "RPM",
            "pH", "TDS", "Turbidity", "Chlorine", "Flow Rate", "Water Level",
            "Motion Detected", "Air Quality", "Occupancy", "Lux",
            "Risk Score", "Status", "Timestamp"
        ]
        self._apply_header_style(ws, headers)

        query = select(SensorReading)
        conditions = []
        if device_id:
            conditions.append(SensorReading.device_id == device_id)
        if module_type:
            conditions.append(SensorReading.module_type == module_type)
        if start_date:
            try:
                sd = datetime.fromisoformat(start_date.replace("Z", "+00:00"))
                conditions.append(SensorReading.created_at >= sd)
            except ValueError:
                pass
        if end_date:
            try:
                ed = datetime.fromisoformat(end_date.replace("Z", "+00:00"))
                conditions.append(SensorReading.created_at <= ed)
            except ValueError:
                pass
        if conditions:
            query = query.where(and_(*conditions))
        query = query.order_by(desc(SensorReading.created_at)).limit(10000)
        result = await self.db.execute(query)
        readings = result.scalars().all()

        severity_fills = {
            "critical": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "warning": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        }

        for row_idx, r in enumerate(readings, 2):
            ws.cell(row=row_idx, column=1, value=r.device_id)
            ws.cell(row=row_idx, column=2, value=r.module_type)
            ws.cell(row=row_idx, column=3, value=r.temperature)
            ws.cell(row=row_idx, column=4, value=r.humidity)
            ws.cell(row=row_idx, column=5, value="Yes" if r.door_open else "No")
            ws.cell(row=row_idx, column=6, value=r.door_open_seconds)
            ws.cell(row=row_idx, column=7, value="Yes" if r.power_available else "No")
            ws.cell(row=row_idx, column=8, value=r.gas_level)
            ws.cell(row=row_idx, column=9, value=r.compressor_current)
            ws.cell(row=row_idx, column=10, value="On" if r.compressor_on else "Off")
            ws.cell(row=row_idx, column=11, value=r.vibration)
            ws.cell(row=row_idx, column=12, value=r.current)
            ws.cell(row=row_idx, column=13, value=r.voltage)
            ws.cell(row=row_idx, column=14, value=r.rpm)
            ws.cell(row=row_idx, column=15, value=r.ph)
            ws.cell(row=row_idx, column=16, value=r.tds)
            ws.cell(row=row_idx, column=17, value=r.turbidity)
            ws.cell(row=row_idx, column=18, value=r.chlorine)
            ws.cell(row=row_idx, column=19, value=r.flow_rate)
            ws.cell(row=row_idx, column=20, value=r.water_level)
            ws.cell(row=row_idx, column=21, value="Yes" if r.motion_detected else "No")
            ws.cell(row=row_idx, column=22, value=r.air_quality)
            ws.cell(row=row_idx, column=23, value=r.occupancy)
            ws.cell(row=row_idx, column=24, value=r.lux)
            ws.cell(row=row_idx, column=25, value=r.risk_score)
            ws.cell(row=row_idx, column=26, value=r.status)
            ws.cell(row=row_idx, column=27, value=r.created_at.strftime("%Y-%m-%d %H:%M:%S") if r.created_at else "")

        self._apply_borders(ws)
        self._auto_column_widths(ws)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    async def export_alerts(self, device_id=None, severity=None, module_type=None):
        wb = Workbook()
        ws = wb.active
        ws.title = "Alerts"

        headers = ["Device ID", "Module Type", "Alert Type", "Severity", "Message", "Acknowledged", "Created At", "Acknowledged At"]
        self._apply_header_style(ws, headers)

        query = select(Alert)
        conditions = []
        if device_id:
            conditions.append(Alert.device_id == device_id)
        if severity:
            conditions.append(Alert.severity == severity)
        if module_type:
            conditions.append(Alert.module_type == module_type)
        if conditions:
            query = query.where(and_(*conditions))
        query = query.order_by(desc(Alert.created_at)).limit(5000)
        result = await self.db.execute(query)
        alerts = result.scalars().all()

        severity_fills = {
            "critical": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "warning": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "info": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        }

        for row_idx, a in enumerate(alerts, 2):
            ws.cell(row=row_idx, column=1, value=a.device_id)
            ws.cell(row=row_idx, column=2, value=a.module_type)
            ws.cell(row=row_idx, column=3, value=a.alert_type)
            sev_cell = ws.cell(row=row_idx, column=4, value=a.severity)
            if a.severity in severity_fills:
                sev_cell.fill = severity_fills[a.severity]
            ws.cell(row=row_idx, column=5, value=a.message)
            ws.cell(row=row_idx, column=6, value="Yes" if a.acknowledged else "No")
            ws.cell(row=row_idx, column=7, value=a.created_at.strftime("%Y-%m-%d %H:%M:%S") if a.created_at else "")
            ws.cell(row=row_idx, column=8, value=a.acknowledged_at.strftime("%Y-%m-%d %H:%M:%S") if a.acknowledged_at else "")

        self._apply_borders(ws)
        self._auto_column_widths(ws)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    async def export_analytics(self, device_id=None, start_date=None, end_date=None, module_type=None):
        wb = Workbook()
        ws = wb.active
        ws.title = "Analytics"

        headers = [
            "Device ID", "Module Type", "Date", "Avg Temp", "Min Temp", "Max Temp",
            "Avg Humidity", "Door Open Count", "Door Open Duration",
            "Compressor Runtime (min)", "Power Failures", "Avg Risk Score",
            "Avg Vibration", "Avg Current", "Avg Voltage", "Avg RPM",
            "Avg pH", "Avg TDS", "Avg Flow Rate", "Avg Water Level",
            "Motion Events", "Avg Air Quality", "Avg Occupancy", "Avg Lux"
        ]
        self._apply_header_style(ws, headers)

        query = select(Analytics)
        conditions = []
        if device_id:
            conditions.append(Analytics.device_id == device_id)
        if module_type:
            conditions.append(Analytics.module_type == module_type)
        if start_date:
            try:
                sd = datetime.fromisoformat(start_date.replace("Z", "+00:00"))
                conditions.append(Analytics.date >= sd)
            except ValueError:
                pass
        if end_date:
            try:
                ed = datetime.fromisoformat(end_date.replace("Z", "+00:00"))
                conditions.append(Analytics.date <= ed)
            except ValueError:
                pass
        if conditions:
            query = query.where(and_(*conditions))
        query = query.order_by(desc(Analytics.date)).limit(5000)
        result = await self.db.execute(query)
        records = result.scalars().all()

        for row_idx, a in enumerate(records, 2):
            ws.cell(row=row_idx, column=1, value=a.device_id)
            ws.cell(row=row_idx, column=2, value=a.module_type)
            ws.cell(row=row_idx, column=3, value=a.date.strftime("%Y-%m-%d") if a.date else "")
            ws.cell(row=row_idx, column=4, value=a.avg_temp)
            ws.cell(row=row_idx, column=5, value=a.min_temp)
            ws.cell(row=row_idx, column=6, value=a.max_temp)
            ws.cell(row=row_idx, column=7, value=a.avg_humidity)
            ws.cell(row=row_idx, column=8, value=a.door_open_count)
            ws.cell(row=row_idx, column=9, value=a.door_open_duration)
            ws.cell(row=row_idx, column=10, value=a.compressor_runtime)
            ws.cell(row=row_idx, column=11, value=a.power_failure_count)
            ws.cell(row=row_idx, column=12, value=a.avg_risk_score)
            ws.cell(row=row_idx, column=13, value=a.avg_vibration)
            ws.cell(row=row_idx, column=14, value=a.avg_current)
            ws.cell(row=row_idx, column=15, value=a.avg_voltage)
            ws.cell(row=row_idx, column=16, value=a.avg_rpm)
            ws.cell(row=row_idx, column=17, value=a.avg_ph)
            ws.cell(row=row_idx, column=18, value=a.avg_tds)
            ws.cell(row=row_idx, column=19, value=a.avg_flow_rate)
            ws.cell(row=row_idx, column=20, value=a.avg_water_level)
            ws.cell(row=row_idx, column=21, value=a.motion_event_count)
            ws.cell(row=row_idx, column=22, value=a.avg_air_quality)
            ws.cell(row=row_idx, column=23, value=a.avg_occupancy)
            ws.cell(row=row_idx, column=24, value=a.avg_lux)

        self._apply_borders(ws)
        self._auto_column_widths(ws)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    async def export_reports(self, report_type=None, module_type=None):
        wb = Workbook()
        ws = wb.active
        ws.title = "Reports"

        headers = ["Device ID", "Module Type", "Report Type", "Report Data", "File Path", "Created At"]
        self._apply_header_style(ws, headers)

        query = select(Report)
        conditions = []
        if report_type:
            conditions.append(Report.report_type == report_type)
        if module_type:
            conditions.append(Report.module_type == module_type)
        if conditions:
            query = query.where(and_(*conditions))
        query = query.order_by(desc(Report.created_at)).limit(5000)
        result = await self.db.execute(query)
        reports = result.scalars().all()

        for row_idx, r in enumerate(reports, 2):
            ws.cell(row=row_idx, column=1, value=r.device_id)
            ws.cell(row=row_idx, column=2, value=r.module_type)
            ws.cell(row=row_idx, column=3, value=r.report_type)
            ws.cell(row=row_idx, column=4, value=str(r.report_data) if r.report_data else "")
            ws.cell(row=row_idx, column=5, value=r.file_path or "")
            ws.cell(row=row_idx, column=6, value=r.created_at.strftime("%Y-%m-%d %H:%M:%S") if r.created_at else "")

        self._apply_borders(ws)
        self._auto_column_widths(ws)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    async def export_device_logs(self, device_id=None, module_type=None):
        wb = Workbook()
        ws = wb.active
        ws.title = "Device Logs"

        headers = ["Device ID", "Zone", "Name", "Module Type", "Group", "Status", "Firmware", "IP Address", "MAC Address", "Last Heartbeat", "Created At"]
        self._apply_header_style(ws, headers)

        query = select(Device)
        conditions = []
        if device_id:
            conditions.append(Device.device_id == device_id)
        if module_type:
            conditions.append(Device.module_type == module_type)
        if conditions:
            query = query.where(and_(*conditions))
        query = query.order_by(desc(Device.created_at))
        result = await self.db.execute(query)
        devices = result.scalars().all()

        for row_idx, d in enumerate(devices, 2):
            ws.cell(row=row_idx, column=1, value=d.device_id)
            ws.cell(row=row_idx, column=2, value=d.zone)
            ws.cell(row=row_idx, column=3, value=d.name)
            ws.cell(row=row_idx, column=4, value=d.module_type)
            ws.cell(row=row_idx, column=5, value=d.group_name)
            status_cell = ws.cell(row=row_idx, column=6, value=d.status)
            if d.status == "active":
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif d.status == "disabled":
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            ws.cell(row=row_idx, column=7, value=d.firmware_version)
            ws.cell(row=row_idx, column=8, value=d.ip_address)
            ws.cell(row=row_idx, column=9, value=d.mac_address)
            ws.cell(row=row_idx, column=10, value=d.last_heartbeat.strftime("%Y-%m-%d %H:%M:%S") if d.last_heartbeat else "")
            ws.cell(row=row_idx, column=11, value=d.created_at.strftime("%Y-%m-%d %H:%M:%S") if d.created_at else "")

        self._apply_borders(ws)
        self._auto_column_widths(ws)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
