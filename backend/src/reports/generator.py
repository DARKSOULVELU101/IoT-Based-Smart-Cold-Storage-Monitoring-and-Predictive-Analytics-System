import io
import json
from datetime import datetime, timedelta
from typing import Optional
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, desc, func
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from src.database.models import SensorReading, Alert, Analytics, Report, Device


class ReportGenerator:
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
    THIN_BORDER = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    MODULE_LABELS = {
        "cold_storage": "Cold Storage",
        "machine_health": "Machine Health",
        "water_quality": "Water Quality",
        "warehouse": "Warehouse",
    }

    def __init__(self, db: AsyncSession):
        self.db = db

    def _style_sheet(self, ws, headers):
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER
        for col in range(1, ws.max_column + 1):
            max_len = max((len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1)), default=10)
            ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 40)

    async def _get_readings(self, device_id, start, end):
        result = await self.db.execute(
            select(SensorReading).where(
                and_(
                    SensorReading.device_id == device_id,
                    SensorReading.created_at >= start,
                    SensorReading.created_at < end,
                )
            ).order_by(SensorReading.created_at)
        )
        return list(result.scalars().all())

    async def _get_alerts(self, device_id, start, end):
        result = await self.db.execute(
            select(Alert).where(
                and_(
                    Alert.device_id == device_id,
                    Alert.created_at >= start,
                    Alert.created_at < end,
                )
            ).order_by(Alert.created_at)
        )
        return list(result.scalars().all())

    def _compute_stats(self, readings):
        if not readings:
            return {}
        temps = [r.temperature for r in readings if r.temperature is not None]
        humids = [r.humidity for r in readings if r.humidity is not None]
        risks = [r.risk_score for r in readings]
        module_type = readings[0].module_type if readings else "cold_storage"

        stats = {
            "module_type": module_type,
            "total_readings": len(readings),
            "avg_risk_score": round(sum(risks) / len(risks), 2) if risks else 0,
            "max_risk_score": max(risks) if risks else 0,
        }

        if temps:
            stats["avg_temp"] = round(sum(temps) / len(temps), 2)
            stats["min_temp"] = round(min(temps), 2)
            stats["max_temp"] = round(max(temps), 2)
        if humids:
            stats["avg_humidity"] = round(sum(humids) / len(humids), 2)

        if module_type == "cold_storage":
            stats["door_open_events"] = sum(1 for r in readings if r.door_open)
            stats["door_open_total_seconds"] = sum(r.door_open_seconds for r in readings if r.door_open_seconds)
            stats["compressor_on_count"] = sum(1 for r in readings if r.compressor_on)
            stats["power_failures"] = sum(1 for r in readings if r.power_available is not None and not r.power_available)
        elif module_type == "machine_health":
            vibrs = [r.vibration for r in readings if r.vibration is not None]
            currs = [r.current for r in readings if r.current is not None]
            volts = [r.voltage for r in readings if r.voltage is not None]
            if vibrs:
                stats["avg_vibration"] = round(sum(vibrs) / len(vibrs), 2)
            if currs:
                stats["avg_current"] = round(sum(currs) / len(currs), 2)
            if volts:
                stats["avg_voltage"] = round(sum(volts) / len(volts), 2)
        elif module_type == "water_quality":
            phs = [r.ph for r in readings if r.ph is not None]
            tdss = [r.tds for r in readings if r.tds is not None]
            if phs:
                stats["avg_ph"] = round(sum(phs) / len(phs), 2)
            if tdss:
                stats["avg_tds"] = round(sum(tdss) / len(tdss), 2)
        elif module_type == "warehouse":
            stats["motion_events"] = sum(1 for r in readings if r.motion_detected)
            aqs = [r.air_quality for r in readings if r.air_quality is not None]
            if aqs:
                stats["avg_air_quality"] = round(sum(aqs) / len(aqs), 2)

        return stats

    async def _save_report(self, device_id, report_type, report_data, module_type="cold_storage"):
        report = Report(
            device_id=device_id,
            module_type=module_type,
            report_type=report_type,
            report_data=report_data,
            file_path=None,
        )
        self.db.add(report)
        await self.db.flush()
        await self.db.refresh(report)
        return report

    async def generate_daily_report(self, device_id, date, module_type="cold_storage"):
        day_start = date.replace(hour=0, minute=0, second=0, microsecond=0)
        day_end = day_start + timedelta(days=1)

        readings = await self._get_readings(device_id, day_start, day_end)
        alerts = await self._get_alerts(device_id, day_start, day_end)
        stats = self._compute_stats(readings)
        label = self.MODULE_LABELS.get(module_type, module_type)

        wb = Workbook()
        ws = wb.active
        ws.title = "Daily Report"

        ws.cell(row=1, column=1, value=f"Daily {label} Report").font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=f"Device: {device_id}")
        ws.cell(row=3, column=1, value=f"Date: {day_start.strftime('%Y-%m-%d')}")

        headers = ["Metric", "Value"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=5, column=col, value=h)
            c.font = self.HEADER_FONT
            c.fill = self.HEADER_FILL

        metrics = [("Total Readings", stats.get("total_readings", 0))]
        if "avg_temp" in stats:
            metrics.extend([
                ("Avg Temperature (C)", stats.get("avg_temp", 0)),
                ("Min Temperature (C)", stats.get("min_temp", 0)),
                ("Max Temperature (C)", stats.get("max_temp", 0)),
            ])
        if "avg_humidity" in stats:
            metrics.append(("Avg Humidity (%)", stats.get("avg_humidity", 0)))
        metrics.extend([
            ("Avg Risk Score", stats.get("avg_risk_score", 0)),
            ("Total Alerts", len(alerts)),
            ("Critical Alerts", sum(1 for a in alerts if a.severity == "critical")),
            ("Warning Alerts", sum(1 for a in alerts if a.severity == "warning")),
        ])

        for row_idx, (metric, value) in enumerate(metrics, 6):
            ws.cell(row=row_idx, column=1, value=metric)
            ws.cell(row=row_idx, column=2, value=value)

        report = await self._save_report(device_id, "daily", stats, module_type)
        return report

    async def generate_weekly_report(self, device_id, week_start, module_type="cold_storage"):
        week_end = week_start + timedelta(days=7)
        readings = await self._get_readings(device_id, week_start, week_end)
        alerts = await self._get_alerts(device_id, week_start, week_end)
        stats = self._compute_stats(readings)
        label = self.MODULE_LABELS.get(module_type, module_type)

        daily_breakdown = []
        for day_offset in range(7):
            day = week_start + timedelta(days=day_offset)
            day_start = day.replace(hour=0, minute=0, second=0, microsecond=0)
            day_end = day_start + timedelta(days=1)
            day_readings = [r for r in readings if day_start <= r.created_at < day_end]
            day_stats = self._compute_stats(day_readings)
            daily_breakdown.append({
                "date": day_start.strftime("%Y-%m-%d"),
                "readings": len(day_readings),
                "avg_temp": day_stats.get("avg_temp"),
                "avg_risk_score": day_stats.get("avg_risk_score", 0),
            })

        wb = Workbook()
        ws = wb.active
        ws.title = "Weekly Report"
        ws.cell(row=1, column=1, value=f"Weekly {label} Report").font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=f"Device: {device_id}")
        ws.cell(row=3, column=1, value=f"Week: {week_start.strftime('%Y-%m-%d')} to {(week_end - timedelta(days=1)).strftime('%Y-%m-%d')}")

        summary_headers = ["Metric", "Value"]
        for col, h in enumerate(summary_headers, 1):
            c = ws.cell(row=5, column=col, value=h)
            c.font = self.HEADER_FONT
            c.fill = self.HEADER_FILL

        metrics = [
            ("Total Readings", stats.get("total_readings", 0)),
            ("Avg Risk Score", stats.get("avg_risk_score", 0)),
            ("Total Alerts", len(alerts)),
            ("Critical Alerts", sum(1 for a in alerts if a.severity == "critical")),
        ]
        for row_idx, (metric, value) in enumerate(metrics, 6):
            ws.cell(row=row_idx, column=1, value=metric)
            ws.cell(row=row_idx, column=2, value=value)

        report_data = {**stats, "daily_breakdown": daily_breakdown}
        report = await self._save_report(device_id, "weekly", report_data, module_type)
        return report

    async def generate_monthly_report(self, device_id, month, year, module_type="cold_storage"):
        month_start = datetime(year, month, 1)
        if month == 12:
            month_end = datetime(year + 1, 1, 1)
        else:
            month_end = datetime(year, month + 1, 1)

        readings = await self._get_readings(device_id, month_start, month_end)
        alerts = await self._get_alerts(device_id, month_start, month_end)
        stats = self._compute_stats(readings)
        label = self.MODULE_LABELS.get(module_type, module_type)

        wb = Workbook()
        ws = wb.active
        ws.title = "Monthly Report"
        ws.cell(row=1, column=1, value=f"Monthly {label} Report").font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=f"Device: {device_id}")
        ws.cell(row=3, column=1, value=f"Month: {month_start.strftime('%B %Y')}")

        headers = ["Metric", "Value"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=5, column=col, value=h)
            c.font = self.HEADER_FONT
            c.fill = self.HEADER_FILL

        metrics = [
            ("Total Readings", stats.get("total_readings", 0)),
            ("Avg Risk Score", stats.get("avg_risk_score", 0)),
            ("Total Alerts", len(alerts)),
            ("Critical Alerts", sum(1 for a in alerts if a.severity == "critical")),
            ("Warning Alerts", sum(1 for a in alerts if a.severity == "warning")),
        ]
        for row_idx, (metric, value) in enumerate(metrics, 6):
            ws.cell(row=row_idx, column=1, value=metric)
            ws.cell(row=row_idx, column=2, value=value)

        report = await self._save_report(device_id, "monthly", stats, module_type)
        return report

    async def generate_compliance_report(self, device_id, start_date, end_date, module_type="cold_storage"):
        readings = await self._get_readings(device_id, start_date, end_date)
        alerts = await self._get_alerts(device_id, start_date, end_date)
        stats = self._compute_stats(readings)
        label = self.MODULE_LABELS.get(module_type, module_type)

        wb = Workbook()
        ws = wb.active
        ws.title = "Compliance Report"
        ws.cell(row=1, column=1, value=f"{label} Compliance Report").font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=f"Device: {device_id}")
        ws.cell(row=3, column=1, value=f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")

        headers = ["Compliance Metric", "Value", "Status"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=5, column=col, value=h)
            c.font = self.HEADER_FONT
            c.fill = self.HEADER_FILL

        total_readings = max(len(readings), 1)

        if module_type == "cold_storage":
            temp_violations = [r for r in readings if r.temperature is not None and (r.temperature < 2 or r.temperature > 8)]
            humidity_violations = [r for r in readings if r.humidity is not None and r.humidity > 70]
            temp_compliance = round((1 - len(temp_violations) / total_readings) * 100, 2)
            humidity_compliance = round((1 - len(humidity_violations) / total_readings) * 100, 2)
            metrics = [
                ("Temperature Compliance (%)", temp_compliance, "PASS" if temp_compliance >= 95 else "FAIL"),
                ("Humidity Compliance (%)", humidity_compliance, "PASS" if humidity_compliance >= 95 else "FAIL"),
                ("Temperature Violations", len(temp_violations), "OK" if len(temp_violations) == 0 else "VIOLATION"),
                ("Humidity Violations", len(humidity_violations), "OK" if len(humidity_violations) == 0 else "VIOLATION"),
            ]
        elif module_type == "machine_health":
            vib_violations = [r for r in readings if r.vibration is not None and r.vibration > 10]
            temp_violations = [r for r in readings if r.temperature is not None and r.temperature > 85]
            vib_compliance = round((1 - len(vib_violations) / total_readings) * 100, 2)
            temp_compliance = round((1 - len(temp_violations) / total_readings) * 100, 2)
            metrics = [
                ("Vibration Compliance (%)", vib_compliance, "PASS" if vib_compliance >= 95 else "FAIL"),
                ("Temperature Compliance (%)", temp_compliance, "PASS" if temp_compliance >= 95 else "FAIL"),
                ("Vibration Violations", len(vib_violations), "OK" if len(vib_violations) == 0 else "VIOLATION"),
                ("Temperature Violations", len(temp_violations), "OK" if len(temp_violations) == 0 else "VIOLATION"),
            ]
        elif module_type == "water_quality":
            ph_violations = [r for r in readings if r.ph is not None and (r.ph < 6.5 or r.ph > 8.5)]
            tds_violations = [r for r in readings if r.tds is not None and r.tds > 500]
            ph_compliance = round((1 - len(ph_violations) / total_readings) * 100, 2)
            tds_compliance = round((1 - len(tds_violations) / total_readings) * 100, 2)
            metrics = [
                ("pH Compliance (%)", ph_compliance, "PASS" if ph_compliance >= 95 else "FAIL"),
                ("TDS Compliance (%)", tds_compliance, "PASS" if tds_compliance >= 95 else "FAIL"),
                ("pH Violations", len(ph_violations), "OK" if len(ph_violations) == 0 else "VIOLATION"),
                ("TDS Violations", len(tds_violations), "OK" if len(tds_violations) == 0 else "VIOLATION"),
            ]
        elif module_type == "warehouse":
            temp_violations = [r for r in readings if r.temperature is not None and r.temperature > 35]
            air_violations = [r for r in readings if r.air_quality is not None and r.air_quality < 50]
            temp_compliance = round((1 - len(temp_violations) / total_readings) * 100, 2)
            air_compliance = round((1 - len(air_violations) / total_readings) * 100, 2)
            metrics = [
                ("Temperature Compliance (%)", temp_compliance, "PASS" if temp_compliance >= 95 else "FAIL"),
                ("Air Quality Compliance (%)", air_compliance, "PASS" if air_compliance >= 95 else "FAIL"),
                ("Temperature Violations", len(temp_violations), "OK" if len(temp_violations) == 0 else "VIOLATION"),
                ("Air Quality Violations", len(air_violations), "OK" if len(air_violations) == 0 else "VIOLATION"),
            ]
        else:
            metrics = []

        metrics.extend([
            ("Critical Alerts", sum(1 for a in alerts if a.severity == "critical"), "OK" if sum(1 for a in alerts if a.severity == "critical") == 0 else "ALERT"),
            ("Avg Risk Score", stats.get("avg_risk_score", 0), "COMPLIANT" if stats.get("avg_risk_score", 0) < 60 else "NON-COMPLIANT"),
        ])

        for row_idx, (metric, value, status_val) in enumerate(metrics, 6):
            ws.cell(row=row_idx, column=1, value=metric)
            ws.cell(row=row_idx, column=2, value=value)
            status_cell = ws.cell(row=row_idx, column=3, value=status_val)
            if status_val in ("PASS", "OK", "COMPLIANT"):
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            else:
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        report_data = {**stats, "total_readings": total_readings}
        report = await self._save_report(device_id, "compliance", report_data, module_type)
        return report

    async def generate_maintenance_report(self, device_id, start_date, end_date, module_type="cold_storage"):
        readings = await self._get_readings(device_id, start_date, end_date)
        stats = self._compute_stats(readings)
        label = self.MODULE_LABELS.get(module_type, module_type)

        wb = Workbook()
        ws = wb.active
        ws.title = "Maintenance Report"
        ws.cell(row=1, column=1, value=f"{label} Maintenance Report").font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=f"Device: {device_id}")
        ws.cell(row=3, column=1, value=f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")

        headers = ["Maintenance Metric", "Value"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=5, column=col, value=h)
            c.font = self.HEADER_FONT
            c.fill = self.HEADER_FILL

        metrics = [
            ("Total Readings", stats.get("total_readings", 0)),
            ("Avg Risk Score", stats.get("avg_risk_score", 0)),
            ("Max Risk Score", stats.get("max_risk_score", 0)),
        ]

        if module_type == "cold_storage":
            compressor_minutes = stats.get("compressor_on_count", 0) * 5
            metrics.extend([
                ("Compressor Runtime (minutes)", compressor_minutes),
                ("Compressor Runtime (hours)", round(compressor_minutes / 60, 2)),
                ("Door Open Events", stats.get("door_open_events", 0)),
                ("Door Open Duration (seconds)", stats.get("door_open_total_seconds", 0)),
                ("Power Failures", stats.get("power_failures", 0)),
            ])
        elif module_type == "machine_health":
            metrics.extend([
                ("Avg Vibration", stats.get("avg_vibration", "N/A")),
                ("Avg Current", stats.get("avg_current", "N/A")),
                ("Avg Voltage", stats.get("avg_voltage", "N/A")),
            ])
        elif module_type == "water_quality":
            metrics.extend([
                ("Avg pH", stats.get("avg_ph", "N/A")),
                ("Avg TDS", stats.get("avg_tds", "N/A")),
            ])
        elif module_type == "warehouse":
            metrics.extend([
                ("Motion Events", stats.get("motion_events", 0)),
                ("Avg Air Quality", stats.get("avg_air_quality", "N/A")),
            ])

        for row_idx, (metric, value) in enumerate(metrics, 6):
            ws.cell(row=row_idx, column=1, value=metric)
            ws.cell(row=row_idx, column=2, value=value)

        report = await self._save_report(device_id, "maintenance", stats, module_type)
        return report

    async def generate_audit_report(self, device_id, start_date, end_date, module_type="cold_storage"):
        readings = await self._get_readings(device_id, start_date, end_date)
        alerts = await self._get_alerts(device_id, start_date, end_date)
        stats = self._compute_stats(readings)
        label = self.MODULE_LABELS.get(module_type, module_type)

        wb = Workbook()
        ws = wb.active
        ws.title = "Audit Report"
        ws.cell(row=1, column=1, value=f"{label} Audit Report").font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=f"Device: {device_id}")
        ws.cell(row=3, column=1, value=f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")

        headers = ["Audit Metric", "Value"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=5, column=col, value=h)
            c.font = self.HEADER_FONT
            c.fill = self.HEADER_FILL

        metrics = [
            ("Total Readings in Period", stats.get("total_readings", 0)),
            ("Total Alerts", len(alerts)),
            ("Critical Alerts", sum(1 for a in alerts if a.severity == "critical")),
            ("Warning Alerts", sum(1 for a in alerts if a.severity == "warning")),
            ("Acknowledged Alerts", sum(1 for a in alerts if a.acknowledged)),
            ("Unacknowledged Alerts", sum(1 for a in alerts if not a.acknowledged)),
            ("Avg Risk Score", stats.get("avg_risk_score", 0)),
            ("Max Risk Score", stats.get("max_risk_score", 0)),
            ("Report Generated", datetime.utcnow().isoformat()),
        ]

        for row_idx, (metric, value) in enumerate(metrics, 6):
            ws.cell(row=row_idx, column=1, value=metric)
            ws.cell(row=row_idx, column=2, value=value)

        report_data = {**stats, "total_alerts": len(alerts)}
        report = await self._save_report(device_id, "audit", report_data, module_type)
        return report
