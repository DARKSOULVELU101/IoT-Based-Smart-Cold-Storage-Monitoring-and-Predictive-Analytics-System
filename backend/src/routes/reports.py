from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional
from datetime import datetime, timedelta
from io import BytesIO
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, desc
from src.database.connection import get_db
from src.database.models import Report
from src.reports.generator import ReportGenerator

router = APIRouter(tags=["Reports"])

VALID_MODULE_TYPES = ("cold_storage", "machine_health", "water_quality", "warehouse")


class ReportGenerateRequest(BaseModel):
    device_id: str
    report_type: str
    module_type: Optional[str] = "cold_storage"
    date: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    month: Optional[int] = None
    year: Optional[int] = None


@router.get("/api/reports")
async def get_reports(
    device_id: Optional[str] = Query(None),
    report_type: Optional[str] = Query(None),
    module_type: Optional[str] = Query(None),
    limit: int = Query(50, ge=1, le=200),
    db: AsyncSession = Depends(get_db),
):
    query = select(Report)
    if device_id:
        query = query.where(Report.device_id == device_id)
    if report_type:
        query = query.where(Report.report_type == report_type)
    if module_type and module_type in VALID_MODULE_TYPES:
        query = query.where(Report.module_type == module_type)
    query = query.order_by(desc(Report.created_at)).limit(limit)
    result = await db.execute(query)
    reports = result.scalars().all()
    return {
        "reports": [
            {
                "id": str(r.id),
                "device_id": r.device_id,
                "module_type": r.module_type,
                "report_type": r.report_type,
                "created_at": r.created_at.isoformat() if r.created_at else None,
            }
            for r in reports
        ],
        "count": len(reports),
    }


@router.post("/api/reports/generate", status_code=status.HTTP_201_CREATED)
async def generate_report(request: ReportGenerateRequest, db: AsyncSession = Depends(get_db)):
    generator = ReportGenerator(db)
    module_type = request.module_type if request.module_type in VALID_MODULE_TYPES else "cold_storage"

    try:
        if request.report_type == "daily":
            date = datetime.fromisoformat(request.date) if request.date else datetime.utcnow()
            report = await generator.generate_daily_report(request.device_id, date, module_type)
        elif request.report_type == "weekly":
            date = datetime.fromisoformat(request.date) if request.date else datetime.utcnow()
            week_start = date - timedelta(days=date.weekday())
            report = await generator.generate_weekly_report(request.device_id, week_start, module_type)
        elif request.report_type == "monthly":
            month = request.month or datetime.utcnow().month
            year = request.year or datetime.utcnow().year
            report = await generator.generate_monthly_report(request.device_id, month, year, module_type)
        elif request.report_type == "compliance":
            start = datetime.fromisoformat(request.start_date) if request.start_date else datetime.utcnow() - timedelta(days=30)
            end = datetime.fromisoformat(request.end_date) if request.end_date else datetime.utcnow()
            report = await generator.generate_compliance_report(request.device_id, start, end, module_type)
        elif request.report_type == "maintenance":
            start = datetime.fromisoformat(request.start_date) if request.start_date else datetime.utcnow() - timedelta(days=30)
            end = datetime.fromisoformat(request.end_date) if request.end_date else datetime.utcnow()
            report = await generator.generate_maintenance_report(request.device_id, start, end, module_type)
        elif request.report_type == "audit":
            start = datetime.fromisoformat(request.start_date) if request.start_date else datetime.utcnow() - timedelta(days=30)
            end = datetime.fromisoformat(request.end_date) if request.end_date else datetime.utcnow()
            report = await generator.generate_audit_report(request.device_id, start, end, module_type)
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Unknown report type: {request.report_type}. Valid: daily, weekly, monthly, compliance, maintenance, audit",
            )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=str(e))

    await db.commit()
    return {
        "id": str(report.id),
        "device_id": report.device_id,
        "module_type": report.module_type,
        "report_type": report.report_type,
        "created_at": report.created_at.isoformat() if report.created_at else None,
        "message": "Report generated successfully",
    }


@router.get("/api/reports/{report_id}")
async def get_report(report_id: str, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Report).where(Report.id == report_id))
    report = result.scalar_one_or_none()
    if not report:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Report not found")
    return {
        "id": str(report.id),
        "device_id": report.device_id,
        "module_type": report.module_type,
        "report_type": report.report_type,
        "report_data": report.report_data,
        "file_path": report.file_path,
        "created_at": report.created_at.isoformat() if report.created_at else None,
    }


@router.get("/api/reports/{report_id}/download")
async def download_report(report_id: str, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Report).where(Report.id == report_id))
    report = result.scalar_one_or_none()
    if not report:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Report not found")

    generator = ReportGenerator(db)
    try:
        if report.report_type == "daily":
            generated = await generator.generate_daily_report(report.device_id, report.created_at or datetime.utcnow(), report.module_type)
        elif report.report_type == "weekly":
            generated = await generator.generate_weekly_report(report.device_id, report.created_at or datetime.utcnow(), report.module_type)
        elif report.report_type == "monthly":
            dt = report.created_at or datetime.utcnow()
            generated = await generator.generate_monthly_report(report.device_id, dt.month, dt.year, report.module_type)
        else:
            start = (report.created_at or datetime.utcnow()) - timedelta(days=30)
            end = report.created_at or datetime.utcnow()
            generated = await generator.generate_compliance_report(report.device_id, start, end, report.module_type)
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=str(e))

    import io
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"{report.report_type.title()} Report"
    if generated.report_data:
        ws.cell(row=1, column=1, value="Report Data")
        row_idx = 2
        for key, value in generated.report_data.items():
            ws.cell(row=row_idx, column=1, value=key)
            ws.cell(row=row_idx, column=2, value=str(value))
            row_idx += 1
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{report.report_type}_{report.module_type}_report_{report.device_id}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
